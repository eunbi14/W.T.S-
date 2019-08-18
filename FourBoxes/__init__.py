from flask import Flask, render_template
from datetime import datetime
from glob import glob

import matplotlib.pyplot as plt
import pandas as pd, numpy as np
import openpyxl


file = glob("4rf_predict.py")
for f in file:
    exec(open(f, encoding='UTF-8').read())

book = openpyxl.load_workbook("total경복궁.xlsx")
sheet9 = book.get_sheet_by_name("2019경복궁")
sheet8 = book.get_sheet_by_name("2018경복궁") # Matirx[7]
sheet7 = book.get_sheet_by_name("2017경복궁") # Matrix[6]
sheet6 = book.get_sheet_by_name("2016경복궁") # Matrix[5]
sheet5 = book.get_sheet_by_name("2015경복궁") # Matrix[4]
sheet4 = book.get_sheet_by_name("2014경복궁") # Matrix[3]
sheet3 = book.get_sheet_by_name("2013경복궁") # Matrix[2]
sheet2 = book.get_sheet_by_name("2012경복궁") # Matrix[1]
sheet1 = book.get_sheet_by_name("2011경복궁") # Matrix[0]

Matrix = [[0]*31 for i in range(12)]
Matrix2012 = [[0]*31 for i in range(12)]
Matrix2013 = [[0]*31 for i in range(12)]
Matrix2014 = [[0]*31 for i in range(12)]
Matrix2015 = [[0]*31 for i in range(12)]
Matrix2016 = [[0]*31 for i in range(12)]
Matrix2017 = [[0]*31 for i in range(12)]
Matrix2018 = [[0]*31 for i in range(12)]
Matrix2019 = [[0]*31 for i in range(12)] 

r = 2
now = datetime.now()
year = now.year

j = 0
for i in range(365):
    month = sheet1.cell(row=r, column=7).value - 1
    month2 = sheet1.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix[month][j] = sheet1.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet2.cell(row=r, column=7).value - 1
    month2 = sheet2.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2012[month][j] = sheet2.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet3.cell(row=r, column=7).value - 1
    month2 = sheet3.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2013[month][j] = sheet3.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet4.cell(row=r, column=7).value - 1
    month2 = sheet4.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2014[month][j] = sheet4.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet5.cell(row=r, column=7).value - 1
    month2 = sheet5.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2015[month][j] = sheet5.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet6.cell(row=r, column=7).value - 1
    month2 = sheet6.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2016[month][j] = sheet6.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet7.cell(row=r, column=7).value - 1
    month2 = sheet7.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2017[month][j] = sheet7.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0
for i in range(365):
    month = sheet8.cell(row=r, column=7).value - 1
    month2 = sheet8.cell(row=r-1, column=7).value
    if i!=0:
        if month != month2 -1:
            j=0
        else:
            j=j+1
    Matrix2018[month][j] = sheet8.cell(row=r,column=3).value
    r = r+1
    
r = 2
j = 0

now = datetime.now()
def f(x):
    return { 1:0, 2:31, 43:59, 4:90, 5:120, 6:151, 7:181, 8:212, 9:243, 10:273, 
            11:304, 12:334}[x]

numday = f(now.month)
if (now.year-2016)%4==0:
    numday = numday +1

numday = numday + now.day

for i in range(numday-1):
    month = sheet9.cell(row = r, column = 7).value - 1
    month2 = sheet9.cell(row = r - 1, column = 7).value
    if i!=0:
        if month != month2 - 1:
            j = 0
        else:
            j = j + 1
    Matrix2019[month][j] = sheet9.cell(row = r, column = 15).value
    r = r + 1
    
app = Flask(__name__)

@app.route('/')
def home():
  return render_template('index.html')

@app.route('/preview_castle.html')
def preview():
    return render_template('preview_castle.html', test=Matrix, 
                           test2012 = Matrix2012, test2013 = Matrix2013, 
                           test2014 = Matrix2014, test2015 = Matrix2015, 
                           test2016 = Matrix2016, test2017 = Matrix2017, 
                           test2018 = Matrix2018, test2019 = Matrix2019)
@app.route('/predict_castle')
def predict_castle():
    return render_template('predict_castle.html')
@app.route('/map_castle.html')
def map_castle():
    return render_template('map_castle.html')
@app.route('/info_castle_1.html')
def info_castle_1():
    return render_template('info_castle_1.html')
@app.route('/info_castle_2.html')
def info_castle_2():
    return render_template('info_castle_2.html')
@app.route('/info_castle_3.html')
def info_castle_3():
    return render_template('info_castle_3.html')

@app.route('/index.html')
def home2():
    return render_template('index.html')
"""
@app.route('/select.html')
def select():
    return render_template('select.html')
    """
