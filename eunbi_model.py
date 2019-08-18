# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:35:24 2019
@author: seoyoung
"""# -*- coding: utf-8 -*-
"""
Created on Sun Jul 21 17:56:44 2019
@author: seoyoung
"""
import matplotlib.pyplot as plt
import pandas as pd, numpy as np
import openpyxl
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import r2_score
import pylab as plot
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.externals import joblib
book = openpyxl.load_workbook("total경복궁.xlsx")

sheet9 = book.get_sheet_by_name("2019경복궁")
sheet8 = book.get_sheet_by_name("2018경복궁") 
sheet7 = book.get_sheet_by_name("2017경복궁")
sheet6 = book.get_sheet_by_name("2016경복궁")
sheet5 = book.get_sheet_by_name("2015경복궁")
sheet4 = book.get_sheet_by_name("2014경복궁")
sheet3 = book.get_sheet_by_name("2013경복궁")
sheet2 = book.get_sheet_by_name("2012경복궁")
sheet1 = book.get_sheet_by_name("2011경복궁")

date=[]
day =[]
cloud=[]
visit=[]
temper=[]
rain=[]
holiday=[]
year = []
culture = []
month = []
night = []
closed = []
    
r=2


r=2

for row in range(1,366):
    #11년도 데이터
    date.append(sheet1.cell(row=r,column=1).value)
    day.append(sheet1.cell(row=r,column=2).value)
    visit.append(sheet1.cell(row=r,column=3).value)

    temper.append(sheet1.cell(row=r,column=4).value)
    holiday.append(sheet1.cell(row=r,column=14).value)
    cloud.append(sheet1.cell(row=r,column=6).value)
    month.append(sheet1.cell(row=r,column=7).value)
    night.append(sheet1.cell(row = r, column = 8).value)
    rain.append(sheet1.cell(row = r, column = 9).value)
    closed.append(sheet1.cell(row = r, column = 10).value)
    culture.append(sheet1.cell(row = r, column = 11).value)
    year.append(sheet1.cell(row = r, column = 12).value)
    r=r+1 
    
r=2

for row in range(1,367):
    #12년도 데이터
    date.append(sheet2.cell(row=r,column=1).value)
    day.append(sheet2.cell(row=r,column=2).value)
    visit.append(sheet2.cell(row=r,column=3).value)

    temper.append(sheet2.cell(row=r,column=4).value)
    holiday.append(sheet2.cell(row=r,column=14).value)
    cloud.append(sheet2.cell(row=r,column=6).value)
    month.append(sheet2.cell(row=r,column=7).value)
    night.append(sheet2.cell(row = r, column = 8).value)
    rain.append(sheet2.cell(row = r, column = 9).value)
    closed.append(sheet2.cell(row = r, column = 10).value)
    culture.append(sheet2.cell(row = r, column = 11).value)
    year.append(sheet2.cell(row = r, column = 12).value)
    r=r+1
    
r=2

for row in range(1,366):
    #13년도 데이터
    date.append(sheet3.cell(row=r,column=1).value)
    day.append(sheet3.cell(row=r,column=2).value)
    visit.append(sheet3.cell(row=r,column=3).value)

    temper.append(sheet3.cell(row=r,column=4).value)
    holiday.append(sheet3.cell(row=r,column=14).value)
    cloud.append(sheet3.cell(row=r,column=6).value)
    month.append(sheet3.cell(row=r,column=7).value)
    night.append(sheet3.cell(row = r, column = 8).value)
    rain.append(sheet3.cell(row = r, column = 9).value)
    closed.append(sheet3.cell(row = r, column = 10).value)
    culture.append(sheet3.cell(row = r, column = 11).value)
    year.append(sheet3.cell(row = r, column = 12).value)
    r=r+1
    
r=2

for row in range(1,366):
    #14년도 데이터
    date.append(sheet4.cell(row=r,column=1).value)
    day.append(sheet4.cell(row=r,column=2).value)
    visit.append(sheet4.cell(row=r,column=3).value)

    temper.append(sheet4.cell(row=r,column=4).value)
    holiday.append(sheet4.cell(row=r,column=14).value)
    cloud.append(sheet4.cell(row=r,column=6).value)
    month.append(sheet4.cell(row=r,column=7).value)
    night.append(sheet4.cell(row = r, column = 8).value)
    rain.append(sheet4.cell(row = r, column = 9).value)
    closed.append(sheet4.cell(row = r, column = 10).value)
    culture.append(sheet4.cell(row = r, column = 11).value)
    year.append(sheet4.cell(row = r, column = 12).value)
    r=r+1

    
r=2

for row in range(1,366):
    #15년도 데이터
    date.append(sheet5.cell(row=r,column=1).value)
    day.append(sheet5.cell(row=r,column=2).value)
    visit.append(sheet5.cell(row=r,column=3).value)

    temper.append(sheet5.cell(row=r,column=4).value)
    holiday.append(sheet5.cell(row=r,column=14).value)
    cloud.append(sheet5.cell(row=r,column=6).value)
    month.append(sheet5.cell(row=r,column=7).value)
    night.append(sheet5.cell(row = r, column = 8).value)
    rain.append(sheet5.cell(row = r, column = 9).value)
    closed.append(sheet5.cell(row = r, column = 10).value)
    culture.append(sheet5.cell(row = r, column = 11).value)
    year.append(sheet5.cell(row = r, column = 12).value)
    r=r+1
    
r=2

for row in range(1,367):
    #16년도 데이터
    date.append(sheet6.cell(row=r,column=1).value)
    day.append(sheet6.cell(row=r,column=2).value)
    visit.append(sheet6.cell(row=r,column=3).value)

    temper.append(sheet6.cell(row=r,column=4).value)
    holiday.append(sheet6.cell(row=r,column=14).value)
    cloud.append(sheet6.cell(row=r,column=6).value)
    month.append(sheet6.cell(row=r,column=7).value)
    night.append(sheet6.cell(row = r, column = 8).value)
    rain.append(sheet6.cell(row = r, column = 9).value)
    closed.append(sheet6.cell(row = r, column = 10).value)
    culture.append(sheet6.cell(row = r, column = 11).value)
    year.append(sheet6.cell(row = r, column = 12).value)
    r=r+1
    
r=2


for row in range(1,366):
    #17년도 데이터
    date.append(sheet7.cell(row=r,column=1).value)
    day.append(sheet7.cell(row=r,column=2).value)
    visit.append(sheet7.cell(row=r,column=3).value)

    temper.append(sheet7.cell(row=r,column=4).value)
    holiday.append(sheet7.cell(row=r,column=14).value)
    cloud.append(sheet7.cell(row=r,column=6).value)
    month.append(sheet7.cell(row=r,column=7).value)
    night.append(sheet7.cell(row = r, column = 8).value)
    rain.append(sheet7.cell(row = r, column = 9).value)
    closed.append(sheet7.cell(row = r, column = 10).value)
    culture.append(sheet7.cell(row = r, column = 11).value)
    year.append(sheet7.cell(row = r, column = 12).value)
    r=r+1
 
r=2

for row in range(1,366):
    #18년도 데이터
    date.append(sheet8.cell(row=r,column=1).value)
    day.append(sheet8.cell(row=r,column=2).value)
    visit.append(sheet8.cell(row=r,column=3).value)

    temper.append(sheet8.cell(row=r,column=4).value)
    holiday.append(sheet8.cell(row=r,column=14).value)
    cloud.append(sheet8.cell(row=r,column=6).value)
    month.append(sheet8.cell(row=r,column=7).value)
    night.append(sheet8.cell(row = r, column = 8).value)
    rain.append(sheet8.cell(row = r, column = 9).value)
    closed.append(sheet8.cell(row = r, column = 10).value)
    culture.append(sheet8.cell(row = r, column = 11).value)
    year.append(sheet8.cell(row = r, column = 12).value)
    r=r+1
    
r = 2
for row in range(1, 224):
    # 19년도 데이터
    date.append(sheet9.cell(row = r, column = 1).value)
    day.append(sheet9.cell(row = r, column = 2).value)
    visit.append(sheet9.cell(row=r,column=3).value)
    temper.append(sheet9.cell(row = r, column = 4).value)
    holiday.append(sheet9.cell(row = r, column = 14).value)
    cloud.append(sheet9.cell(row = r, column = 6).value)
    month.append(sheet9.cell(row = r, column = 7).value)
    night.append(sheet9.cell(row = r, column = 8).value)
    rain.append(sheet9.cell(row = r, column = 9).value)
    closed.append(sheet9.cell(row = r, column = 10).value)
    culture.append(sheet9.cell(row = r, column = 11).value)
    year.append(sheet9.cell(row = r, column = 12).value)
    r = r + 1
    
temper=np.array(temper)
temper_mean=np.mean(temper)
temper-=temper_mean
temper_std=np.std(temper)
temper /=temper_std

cloud=np.array(cloud)
cloud_mean=np.mean(cloud)
cloud = cloud - cloud_mean
cloud_std=np.std(cloud)
cloud /=cloud_std


rain=np.array(rain)
rain_mean=np.mean(rain)
rain = rain - rain_mean
rain_std=np.std(rain)
rain /=rain_std
"""
print("temper_mean:",temper_mean)
print("temper_std:",temper_std)
print("cloud_mean:",cloud_mean)
print("cloud_std:",cloud_std)
print("rain_mean:",rain_mean)
print("rain_std:",rain_std)
"""
a = int(1)
b = int (0)
day_class = {
        "월":[a,b,b,b,b,b,b],
        "화":[b,a,b,b,b,b,b],
        "수":[b,b,a,b,b,b,b],
        "목":[b,b,b,a,b,b,b],
        "금":[b,b,b,b,a,b,b],
        "토":[b,b,b,b,b,a,b],
        "일":[b,b,b,b,b,b,a]
        }


month_class = {
        1:[a,b,b,b,b,b,b,b,b,b,b,b],
        2:[b,a,b,b,b,b,b,b,b,b,b,b],
        3:[b,b,a,b,b,b,b,b,b,b,b,b],
        4:[b,b,b,a,b,b,b,b,b,b,b,b],
        5:[b,b,b,b,a,b,b,b,b,b,b,b],
        6:[b,b,b,b,b,a,b,b,b,b,b,b],
        7:[b,b,b,b,b,b,a,b,b,b,b,b],
        8:[b,b,b,b,b,b,b,a,b,b,b,b],
        9:[b,b,b,b,b,b,b,b,a,b,b,b],
        10:[b,b,b,b,b,b,b,b,b,a,b,b],
        11:[b,b,b,b,b,b,b,b,b,b,a,b],
        12:[b,b,b,b,b,b,b,b,b,b,b,a]
        }

year_class = {
        2011 : [a, b, b, b, b, b, b, b, b],
        2012 : [b, a, b, b, b, b, b, b, b],
        2013 : [b, b, a, b, b, b, b, b, b],
        2014 : [b, b, b, a, b, b, b, b, b],
        2015 : [b, b, b, b, a, b, b, b, b],
        2016 : [b, b, b, b, b, a, b, b, b],
        2017 : [b, b, b, b, b, b, a, b, b],
        2018 : [b, b, b, b, b, b, b, a, b],
        2019 : [b, b, b, b, b, b, b, b, a]
        }

w=np.empty((len(day),7),dtype="int")
#isit = visit.as_matrix()
m=np.empty((len(month),12),dtype="int")
y = np.empty((len(year), 9), dtype = "int")

for i , v in enumerate(day):
    w[i] = day_class[v]

for i , v in enumerate(month):
    m[i] = month_class[v]
    
for i, v in enumerate(year):
    y[i] = year_class[v]

x1=np.zeros((w.shape[0],w.shape[1]+1))
x1[:,:-1]=w
x1[:,-1]=temper


x2=np.zeros((x1.shape[0],x1.shape[1]+1))
x2[:,:-1]=x1
x2[:,-1]= holiday

x3=np.zeros((x2.shape[0],x2.shape[1]+1))
x3[:,:-1]=x2
x3[:,-1]= cloud

x4=np.hstack((x3,m))

x5 = np.zeros((x4.shape[0], x4.shape[1] + 1))
x5[:, :-1] = x4
x5[:, -1] = night

x6 = np.zeros((x5.shape[0], x5.shape[1] + 1))
x6[:, :-1] = x5
x6[:, -1] = rain

x7 = np.zeros((x6.shape[0], x6.shape[1] + 1))
x7[:, :-1] = x6
x7[:, -1] = closed

x8 = np.zeros((x7.shape[0], x7.shape[1] + 1))
x8[:, :-1] = x7
x8[:, -1] = culture

x9=np.hstack((x8,y))

x=x9
y=visit


x_train = x[:-223]
y_train = y[:-223]
x_test = x[-223:]
y_test = y[-223:]

forest = RandomForestRegressor(n_estimators=2000,criterion='mse',max_depth=30,random_state=4,n_jobs=-1,verbose=1)
#score = cross_val_score(forest,x_train,y_train,cv=k_fold)
#score=score.mean()

#print(score)
forest.fit(x_train,y_train)
#모델을 만듭니다.
r2=mean_absolute_error(y_train,forest.predict(x_train))
print("mae for x_test: ")
print(r2)
print(forest.predict(x_test))

job = "eunbi_model.pkl"


with open(job, 'wb') as file:
    joblib.dump(forest,job)

with open(job, 'rb') as file:
    model = joblib.load(job)
