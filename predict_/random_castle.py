# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:35:24 2019

@author: seoyoung
"""# -*- coding: utf-8 -*-
"""
Created on Sun Jul 21 17:56:44 2019

@author: seoyoung
"""
import matplotlib.pyplot as plt
import pandas as pd, numpy as np
import openpyxl
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import r2_score
import pylab as plot
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.externals import joblib
book = openpyxl.load_workbook("castle_total_closedno.xlsx")


sheet8 = book.get_sheet_by_name("2018경복궁") 
sheet7 = book.get_sheet_by_name("2017경복궁")
sheet6 = book.get_sheet_by_name("2016경복궁")
sheet5 = book.get_sheet_by_name("2015경복궁")
sheet4 = book.get_sheet_by_name("2014경복궁")
sheet3 = book.get_sheet_by_name("2013경복궁")
sheet2 = book.get_sheet_by_name("2012경복궁")
sheet1 = book.get_sheet_by_name("2011경복궁")

date=[]
week=[]
visit=[]
temper=[]
holiday=[]
cloud=[]
month=[]
night=[]
rain=[]
closed=[]
culture=[]
year=[]

r=2

for row in range(1,317):
    #11년도 데이터
    date.append(sheet1.cell(row=r,column=1).value)
    week.append(sheet1.cell(row=r,column=2).value)
    visit.append(sheet1.cell(row=r,column=3).value)
    temper.append(sheet1.cell(row=r,column=4).value)
    holiday.append(sheet1.cell(row=r,column=5).value)
    cloud.append(sheet1.cell(row=r,column=6).value)
    month.append(sheet1.cell(row=r,column=7).value)
    night.append(sheet1.cell(row=r,column=8).value)
    rain.append(sheet1.cell(row=r,column=9).value)
    closed.append(sheet1.cell(row=r,column=10).value)
    year.append(0.1)
    culture.append(sheet1.cell(row=r,column=11).value)
    r=r+1 
    
r=2
for row in range(1,318):
    #12년도 데이터
    date.append(sheet2.cell(row=r,column=1).value)
    week.append(sheet2.cell(row=r,column=2).value)
    visit.append(sheet2.cell(row=r,column=3).value)
    temper.append(sheet2.cell(row=r,column=4).value)
    holiday.append(sheet2.cell(row=r,column=5).value)
    cloud.append(sheet2.cell(row=r,column=6).value)
    month.append(sheet2.cell(row=r,column=7).value)
    night.append(sheet2.cell(row=r,column=8).value)
    rain.append(sheet2.cell(row=r,column=9).value)
    closed.append(sheet2.cell(row=r,column=10).value)
    year.append(0.2)
    culture.append(sheet2.cell(row=r,column=11).value)
    r=r+1
    
r=2

for row in range(1,313):
    #13년도 데이터
    date.append(sheet3.cell(row=r,column=1).value)
    week.append(sheet3.cell(row=r,column=2).value)
    visit.append(sheet3.cell(row=r,column=3).value)
    temper.append(sheet3.cell(row=r,column=4).value)
    holiday.append(sheet3.cell(row=r,column=5).value)
    cloud.append(sheet3.cell(row=r,column=6).value)
    month.append(sheet3.cell(row=r,column=7).value)
    night.append(sheet3.cell(row=r,column=8).value)
    rain.append(sheet3.cell(row=r,column=9).value)
    closed.append(sheet3.cell(row=r,column=10).value)
    year.append(0.3)
    culture.append(sheet3.cell(row=r,column=11).value)
    r=r+1
    
r=2

for row in range(1,315):
    #14년도 데이터
    date.append(sheet4.cell(row=r,column=1).value)
    week.append(sheet4.cell(row=r,column=2).value)
    visit.append(sheet4.cell(row=r,column=3).value)
    temper.append(sheet4.cell(row=r,column=4).value)
    holiday.append(sheet4.cell(row=r,column=5).value)
    cloud.append(sheet4.cell(row=r,column=6).value)
    month.append(sheet4.cell(row=r,column=7).value)
    night.append(sheet4.cell(row=r,column=8).value)
    rain.append(sheet4.cell(row=r,column=9).value)
    closed.append(sheet4.cell(row=r,column=10).value)
    year.append(0.4)
    culture.append(sheet4.cell(row=r,column=11).value)
    r=r+1

    
r=2

for row in range(1,317):
    #15년도 데이터
    date.append(sheet5.cell(row=r,column=1).value)
    week.append(sheet5.cell(row=r,column=2).value)
    visit.append(sheet5.cell(row=r,column=3).value)
    temper.append(sheet5.cell(row=r,column=4).value)
    holiday.append(sheet5.cell(row=r,column=5).value)
    cloud.append(sheet5.cell(row=r,column=6).value)
    month.append(sheet5.cell(row=r,column=7).value)
    night.append(sheet5.cell(row=r,column=8).value)
    rain.append(sheet5.cell(row=r,column=9).value)
    closed.append(sheet5.cell(row=r,column=10).value)
    year.append(0.5)
    culture.append(sheet5.cell(row=r,column=11).value)
    r=r+1
    
r=2

for row in range(1,318):
    #16년도 데이터
    date.append(sheet6.cell(row=r,column=1).value)
    week.append(sheet6.cell(row=r,column=2).value)
    visit.append(sheet6.cell(row=r,column=3).value)
    temper.append(sheet6.cell(row=r,column=4).value)
    holiday.append(sheet6.cell(row=r,column=5).value)
    cloud.append(sheet6.cell(row=r,column=6).value)
    month.append(sheet6.cell(row=r,column=7).value)
    night.append(sheet6.cell(row=r,column=8).value)
    rain.append(sheet6.cell(row=r,column=9).value)
    closed.append(sheet6.cell(row=r,column=10).value)
    year.append(0.6)
    culture.append(sheet6.cell(row=r,column=11).value)
    r=r+1
    
r=2


for row in range(1,318):
    #17년도 데이터
    date.append(sheet7.cell(row=r,column=1).value)
    week.append(sheet7.cell(row=r,column=2).value)
    visit.append(sheet7.cell(row=r,column=3).value)
    temper.append(sheet7.cell(row=r,column=4).value)
    holiday.append(sheet7.cell(row=r,column=5).value)
    cloud.append(sheet7.cell(row=r,column=6).value)
    month.append(sheet7.cell(row=r,column=7).value)
    night.append(sheet7.cell(row=r,column=8).value)
    rain.append(sheet7.cell(row=r,column=9).value)
    closed.append(sheet7.cell(row=r,column=10).value)
    year.append(0.7)
    culture.append(sheet7.cell(row=r,column=11).value)
    r=r+1
 
r=2

for row in range(1,319):
    #18년도 데이터
    date.append(sheet8.cell(row=r,column=1).value)
    week.append(sheet8.cell(row=r,column=2).value)
    visit.append(sheet8.cell(row=r,column=3).value)
    temper.append(sheet8.cell(row=r,column=4).value)
    holiday.append(sheet8.cell(row=r,column=5).value)
    cloud.append(sheet8.cell(row=r,column=6).value)
    month.append(sheet8.cell(row=r,column=7).value)
    night.append(sheet8.cell(row=r,column=8).value)
    rain.append(sheet8.cell(row=r,column=9).value)
    closed.append(sheet8.cell(row=r,column=10).value)
    year.append(0.8)
    culture.append(sheet8.cell(row=r,column=11).value)
    r=r+1

#print(temper)

temper=np.array(temper)
temper_mean=np.mean(temper)
temper-=temper_mean
temper_std=np.std(temper)
temper /=temper_std

cloud=np.array(cloud)
cloud_mean=np.mean(cloud)
cloud-=cloud_mean
cloud_std=np.std(cloud)
cloud /=cloud_std

rain=np.array(rain)
rain_mean=np.mean(rain)
rain-=rain_mean
rain_std=np.std(rain)
rain /=rain_std



a = int(1)
b = int (0)
week_class = {
        "월":[a,b,b,b,b,b,b],
        "화":[b,a,b,b,b,b,b],
        "수":[b,b,a,b,b,b,b],
        "목":[b,b,b,a,b,b,b],
        "금":[b,b,b,b,a,b,b],
        "토":[b,b,b,b,b,a,b],
        "일":[b,b,b,b,b,b,a]
        }



month_class = {
        1:[a,b,b,b,b,b,b,b,b,b,b,b],
        2:[b,a,b,b,b,b,b,b,b,b,b,b],
        3:[b,b,a,b,b,b,b,b,b,b,b,b],
        4:[b,b,b,a,b,b,b,b,b,b,b,b],
        5:[b,b,b,b,a,b,b,b,b,b,b,b],
        6:[b,b,b,b,b,a,b,b,b,b,b,b],
        7:[b,b,b,b,b,b,a,b,b,b,b,b],
        8:[b,b,b,b,b,b,b,a,b,b,b,b],
        9:[b,b,b,b,b,b,b,b,a,b,b,b],
        10:[b,b,b,b,b,b,b,b,b,a,b,b],
        11:[b,b,b,b,b,b,b,b,b,b,a,b],
        12:[b,b,b,b,b,b,b,b,b,b,b,a]
        }


holiday_class = {
        0:0,
        "새해":1,
        "설날":1.5,
        "설날당일":2,
        "삼일":0.5,
        "어린":2,
        "석가":1,
        "현충":1,
        "광복":1,
        "추석":1.5,
        "추석당일":2,
        "대체":1,
        "개천":1,
        "선거":1,
        "연휴":0.5,
        "성탄":1,
        "한글":0.5
        }

w=np.empty((len(week),7),dtype="int")
#isit = visit.as_matrix()
m=np.empty((len(week),12),dtype="int")
h=np.empty((len(week)))

#yy=np.empty((len(week),8),dtype="int")

for i , v in enumerate(week):
    w[i] = week_class[v]

for i , v in enumerate(month):
    m[i] = month_class[v]

for i , v in enumerate(holiday):
    h[i] = holiday_class[v]

h=np.array(h)
h_mean=np.mean(h)
h-=h_mean
h_std=np.std(h)
h /=h_std

x1=np.zeros((w.shape[0],w.shape[1]+1))
x1[:,:-1]=w
x1[:,-1]=temper


x2=np.zeros((x1.shape[0],x1.shape[1]+1))
x2[:,:-1]=x1
x2[:,-1]=rain

x3=np.zeros((x2.shape[0],x2.shape[1]+1))
x3[:,:-1]=x2
x3[:,-1]=culture


x4=np.zeros((x3.shape[0],x3.shape[1]+1))
x4[:,:-1]=x3
x4[:,-1]=cloud

x5=np.hstack((x4,m))

x6=np.zeros((x5.shape[0],x5.shape[1]+1))
x6[:,:-1]=x5
x6[:,-1]=year

x7=np.zeros((x6.shape[0],x6.shape[1]+1))
x7[:,:-1]=x6
x7[:,-1]=night

""" 
x8=np.zeros((x7.shape[0],x7.shape[1]+1))
x8[:,:-1]=x7
x8[:,-1]=closed
"""

x9=np.zeros((x7.shape[0],x7.shape[1]+1))
x9[:,:-1]=x7
x9[:,-1]= h

x=x9
#x=np.hstack((x9,date))


y=visit


num=2200
x_train, y_train = x[:num,], visit[:num]

y_test = visit[num:]
x_test=x[num:]

#x_date=x[num:,28]
#x_train, x_test, y_train,y_test = train_test_split(x,y,test_size=0.1,random_state=1)

#k_fold = KFold(n_splits=10,shuffle=True,random_state=0)

forest = RandomForestRegressor(n_estimators=2000,criterion='mse',max_depth=30,random_state=4,n_jobs=-1,verbose=1)
#score = cross_val_score(forest,x_train,y_train,cv=k_fold)
#score=score.mean()

#print(score)
forest.fit(x_train,y_train)
#모델을 만듭니다.
r2=mean_absolute_error(y_train,forest.predict(x_train))
print("r2???????????????????????????????")

r3=mean_absolute_error(y_test,forest.predict(x_test))
print(r3)

job = "model.pkl"


with open(job, 'wb') as file:
    joblib.dump(forest,job)

with open(job, 'rb') as file:
    model = joblib.load(job)

#score9 = model.score(x_test,y_test)

#print("Test score : {0:.2f} %".format(100*score9))

y_predict = model.predict(x_test)
print(y_predict)
"""
print("????")

 
q = int(len(y_test))

p = int(len(y_train))

literate=1

j=0
j1=int(p/literate)

for i in range(literate):
    
    plt.plot(y_train[j1*j:j1*(j+1)])
    plt.plot(forest.predict(x_train[j1*j:j1*(j+1)]))
    j=j+1
    plt.title('predict')
    plt.ylabel('visit')
    plt.legend(['real','predict'],loc='upper left')
    plt.show()


i=0
y_predict = forest.predict(x[:,:28])
count =0


for i in range(q):
    if(y_test[i]-y_predict[i]>5000 or y_test[i]-y_predict[i]<-5000):
        print("x_test:")
#        print(x_test[i])
        weeek=x_test[i,:7]
        moonth=x_test[i,11:23]
        yeear=x_test[i,23]
        hohoho=x_test[i,25]
        print("week")
        print(weeek)
        print("moonth")
        co=1
        for j in moonth:
            if(j==1):
                print(co)
            co=co+1
        
        print("year:")
        print(yeear*10+2010)
        print("y_test:")
        print(y_test[i])
        print("y_predict:")
        print(y_predict[i])
        count=count+1
    

j=0
j1=int(q/literate)

for i in range(literate):
    
    plt.plot(y_test[j1*j:j1*(j+1)])
    plt.plot(forest.predict(x_test[j1*j:j1*(j+1)]))
    j=j+1
    plt.title('predict')
    plt.ylabel('visit')
    plt.legend(['real','predict'],loc='upper left')
    plt.show()

"""