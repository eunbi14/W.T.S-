# -*- coding: utf-8 -*-
"""
Created on Tue Aug  6 03:04:34 2019

@author: seoyoung
"""
import openpyxl
from sklearn.externals import joblib
import numpy as np
from datetime import datetime
from datetime import timedelta

from selenium import webdriver
#날씨
DRIVER_DIR ='C:\chromedriver'
driver=webdriver.Chrome(DRIVER_DIR)
url="http://www.weatheri.co.kr/forecast/forecast01.php?rid=0201010202&k=1&a_name=%EA%B3%BC%EC%B2%9C"
driver.get(url)
driver.implicitly_wait(5) 

#모델
job = "model.pkl"

with open(job, 'rb') as file:
    model = joblib.load(job)

#holiday,night,culture 데이터
book = openpyxl.load_workbook("read_predict_data.xlsx")
sheet = book.get_sheet_by_name("Sheet1") 


week=[]
visit=[]
temper=[]
holiday=[]
cloud=[]
month=[]
night=[]
rain=[]
closed=[]
culture=[]
year=[]
w=[]
m=[]
h=[]

a = int(1)
b = int (0)
week_class = {
        "월":[a,b,b,b,b,b,b],
        "화":[b,a,b,b,b,b,b],
        "수":[b,b,a,b,b,b,b],
        "목":[b,b,b,a,b,b,b],
        "금":[b,b,b,b,a,b,b],
        "토":[b,b,b,b,b,a,b],
        "일":[b,b,b,b,b,b,a]
        }



month_class = {
        1:[a,b,b,b,b,b,b,b,b,b,b,b],
        2:[b,a,b,b,b,b,b,b,b,b,b,b],
        3:[b,b,a,b,b,b,b,b,b,b,b,b],
        4:[b,b,b,a,b,b,b,b,b,b,b,b],
        5:[b,b,b,b,a,b,b,b,b,b,b,b],
        6:[b,b,b,b,b,a,b,b,b,b,b,b],
        7:[b,b,b,b,b,b,a,b,b,b,b,b],
        8:[b,b,b,b,b,b,b,a,b,b,b,b],
        9:[b,b,b,b,b,b,b,b,a,b,b,b],
        10:[b,b,b,b,b,b,b,b,b,a,b,b],
        11:[b,b,b,b,b,b,b,b,b,b,a,b],
        12:[b,b,b,b,b,b,b,b,b,b,b,a]
        }


holiday_class = {
        0:0,
        "새해":1,
        "설날":1.5,
        "설날당일":2,
        "삼일":0.5,
        "어린":2,
        "석가":1,
        "현충":1,
        "광복":1,
        "추석":1.5,
        "추석당일":2,
        "대체":1,
        "개천":1,
        "선거":1,
        "연휴":0.5,
        "성탄":1,
        "한글":0.5
        }
"""
순서
1.요일
2.기온
3.비
4.문화가있는날
5.cloud
6.month
7.year
8.night
9.holiday

"""


#date[0] ="2018.09.08"
t = ['월','화','수','목','금','토','일']
day1 = datetime.now()
day2 = day1+timedelta(days=1)
day3 = day1+timedelta(days=2)
day4 = day1+timedelta(days=3)
day5 = day1+timedelta(days=4)
day6 = day1+timedelta(days=5)
day7 = day1+timedelta(days=6)
day8 = day1+timedelta(days=7)
day9 = day1+timedelta(days=8)
day10 = day1+timedelta(days=9)

week.append(t[day1.weekday()])
week.append(t[day2.weekday()])
week.append(t[day3.weekday()])
week.append(t[day4.weekday()])
week.append(t[day5.weekday()])
week.append(t[day6.weekday()])
week.append(t[day7.weekday()])
week.append(t[day8.weekday()])
week.append(t[day9.weekday()])
week.append(t[day10.weekday()])

month.append(day1.month)
month.append(day2.month)
month.append(day3.month)
month.append(day4.month)
month.append(day5.month)
month.append(day6.month)
month.append(day7.month)
month.append(day8.month)
month.append(day9.month)
month.append(day10.month)

year.append(day1.year)
year.append(day2.year)
year.append(day3.year)
year.append(day4.year)
year.append(day5.year)
year.append(day6.year)
year.append(day7.year)
year.append(day8.year)
year.append(day9.year)
year.append(day10.year)

year=np.array(year)
year = (year-2010)/10


time1 = datetime(2019,8,1,12,1,1)

r=2+(day1-time1).days

print(sheet.cell(row=r,column=1).value)

for row in range(10):
    holiday.append(sheet.cell(row=r,column=2).value)
    night.append(sheet.cell(row=r,column=3).value)
    culture.append(sheet.cell(row=r,column=4).value)
    r=r+1 
    

w=np.empty((len(week),7),dtype="int")
#isit = visit.as_matrix()
m=np.empty((len(week),12),dtype="int")
h=np.empty(len(week))

for i , v in enumerate(week):
    w[i] = week_class[v]

for i , v in enumerate(month):
    m[i] = month_class[v]


for i , v in enumerate(holiday):
    h[i] = holiday_class[v]

print("week:")
print(w)
print("month:")
print(m)
print("year:")
print(year)
print("holiday:")
print(h)
print("night:")
print(night)
print("culture:")
print(culture)



img_path={
        'http://www.weatheri.co.kr/images/icon_2013_01/01.png':1,
        'http://www.weatheri.co.kr/images/icon_2013_01/02.png':2,
        'http://www.weatheri.co.kr/images/icon_2013_01/18.png':3,
        'http://www.weatheri.co.kr/images/icon_2013_01/21.png':4,
        'http://www.weatheri.co.kr/images/icon_2013_01/03.png':6,
        'http://www.weatheri.co.kr/images/icon_2013_01/13.png':7,
        'http://www.weatheri.co.kr/images/icon_2013_01/07.png':8,
        'http://www.weatheri.co.kr/images/icon_2013_01/04.png':10
        }

day1=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[1]/b")
rain1=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[1]/table/tbody/tr[4]/td[2]/font")
cloud1=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[1]/img")
temp_max1=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/b/font")
temp_min1=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud1.get_attribute('src')
cl1=img_path[cloud_]
print(rain1.text,cl1,temp_max1.text,temp_min1.text)
temper.append((float(temp_max1.text[0:len(temp_max1.text)-2])+float(temp_min1.text[0:len(temp_min1.text)-2]))/2)
if rain1.text[0:len(rain1.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain1.text[0:len(rain1.text)-3]))
cloud.append(float(cl1))

day2=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[2]/b")
rain2=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[2]/table/tbody/tr[4]/td[2]/font")
cloud2=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[1]/img")
temp_max2=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[2]/b/font")
temp_min2=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud2.get_attribute('src')
cl2=img_path[cloud_]
print(day2.text,rain2.text,cl2,temp_max2.text,temp_min2.text)

if rain2.text[0:len(rain2.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain2.text[0:len(rain2.text)-3]))
cloud.append(float(cl2))

temper.append((float(temp_max2.text[0:len(temp_max2.text)-2])+float(temp_min2.text[0:len(temp_min2.text)-2]))/2)

day3=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[3]/b")
rain3=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[3]/table/tbody/tr[4]/td[2]/font")
cloud3=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[1]/img")
temp_max3=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[2]/b/font")
temp_min3=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud3.get_attribute('src')
cl3=img_path[cloud_]
print(day3.text,rain3.text,cl3,temp_max3.text,temp_min3.text)
cloud.append(float(cl3))

if rain3.text[0:len(rain3.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain3.text[0:len(rain3.text)-3]))

temper.append((float(temp_max3.text[0:len(temp_max3.text)-2])+float(temp_min3.text[0:len(temp_min3.text)-2]))/2)
    
day4=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[4]/b")
rain4=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[4]/table/tbody/tr[4]/td[2]/font")
cloud4=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[1]/img")
temp_max4=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[2]/b/font")
temp_min4=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud4.get_attribute('src')
cl4=img_path[cloud_]
print(day4.text,rain4.text,cl4,temp_max4.text,temp_min4.text)
cloud.append(float(cl4))
if rain4.text[0:len(rain4.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain4.text[0:len(rain4.text)-3]))
    
temper.append((float(temp_max4.text[0:len(temp_max4.text)-2])+float(temp_min4.text[0:len(temp_min4.text)-2]))/2)

day5=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[1]/td[5]/b")
rain5=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[5]/table/tbody/tr[4]/td[2]/font")
cloud5=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[1]/img")
temp_max5=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[2]/b/font")
temp_min5=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[2]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud5.get_attribute('src')
cl5=img_path[cloud_]
print(day5.text,rain5.text,cl5,temp_max5.text,temp_min5.text)
cloud.append(float(cl5))
if rain5.text[0:len(rain5.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain5.text[0:len(rain5.text)-3]))

temper.append((float(temp_max5.text[0:len(temp_max5.text)-2])+float(temp_min5.text[0:len(temp_min5.text)-2]))/2)

day6=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[1]/td[1]/b")
rain6=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[1]/table/tbody/tr[4]/td[2]/font")
cloud6=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[1]/img")
temp_max6=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/b/font")
temp_min6=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[1]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud6.get_attribute('src')
cl6=img_path[cloud_]
print(day6.text,rain6.text,cl6,temp_max6.text,temp_min6.text)
cloud.append(float(cl6))
if rain6.text[0:len(rain6.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain6.text[0:len(rain6.text)-3]))

temper.append((float(temp_max6.text[0:len(temp_max6.text)-2])+float(temp_min6.text[0:len(temp_min6.text)-2]))/2)

day7=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[1]/td[2]/b")
rain7=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[2]/table/tbody/tr[4]/td[2]/font")
cloud7=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[1]/img")
temp_max7=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[2]/b/font")
temp_min7=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[2]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud7.get_attribute('src')
cl7=img_path[cloud_]
print(day7.text,rain7.text,cl7,temp_max7.text,temp_min7.text)
cloud.append(float(cl7))
if rain7.text[0:len(rain7.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain7.text[0:len(rain7.text)-3]))

temper.append((float(temp_max7.text[0:len(temp_max7.text)-2])+float(temp_min7.text[0:len(temp_min7.text)-2]))/2)

day8=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[1]/td[3]/b")
rain8=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[3]/table/tbody/tr[4]/td[2]/font")
cloud8=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[1]/img")
temp_max8=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[2]/b/font")
temp_min8=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[3]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud8.get_attribute('src')
cl8=img_path[cloud_]
print(day8.text,rain8.text,cl8,temp_max8.text,temp_min8.text)
cloud.append(float(cl8))
if rain8.text[0:len(rain8.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain8.text[0:len(rain8.text)-3]))

temper.append((float(temp_max8.text[0:len(temp_max8.text)-2])+float(temp_min8.text[0:len(temp_min8.text)-2]))/2)

day9=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[1]/td[4]/b")
rain9=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[4]/table/tbody/tr[4]/td[2]/font")
cloud9=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[1]/img")
temp_max9=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[2]/b/font")
temp_min9=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[4]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud9.get_attribute('src')
cl9=img_path[cloud_]
print(day9.text,rain9.text,cl9,temp_max9.text,temp_min9.text)
cloud.append(float(cl9))
if rain9.text[0:len(rain9.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain9.text[0:len(rain9.text)-3]))

temper.append((float(temp_max9.text[0:len(temp_max9.text)-2])+float(temp_min9.text[0:len(temp_min9.text)-2]))/2)

day10=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[1]/td[5]/b")
rain10=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[5]/table/tbody/tr[4]/td[2]/font")
cloud10=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[1]/img")
temp_max10=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[2]/b/font")
temp_min10=driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[3]/td/table[8]/tbody/tr[2]/td[5]/table/tbody/tr[2]/td[2]/b/b/font")
cloud_=cloud10.get_attribute('src')
cl10=img_path[cloud_]
print(day10.text,rain10.text,cl10,temp_max10.text,temp_min10.text)
cloud.append(float(cl10))
if rain10.text[0:len(rain10.text)-3] == '-':
    rain.append(0)
else: 
    rain.append(float(rain10.text[0:len(rain10.text)-3]))

temper.append((float(temp_max10.text[0:len(temp_max10.text)-2])+float(temp_min10.text[0:len(temp_min10.text)-2]))/2)

print(rain)
print(cloud)
print(temper)
driver.close()

temper_mean = 11.9
temper_std = 10.8

rain_mean = 3.4
rain_std = 13.93

cloud_mean = 4.77
cloud_std = 3.12

temper=np.array(temper)
temper-=temper_mean
temper /=temper_std

cloud=np.array(cloud)
cloud-=cloud_mean
cloud /=cloud_std

rain=np.array(rain)
rain-=rain_mean
rain /=rain_std


x1=np.zeros((w.shape[0],w.shape[1]+1))
x1[:,:-1]=w
x1[:,-1]=temper

x2=np.zeros((x1.shape[0],x1.shape[1]+1))
x2[:,:-1]=x1
x2[:,-1]=rain

x3=np.zeros((x2.shape[0],x2.shape[1]+1))
x3[:,:-1]=x2
x3[:,-1]=culture

x4=np.zeros((x3.shape[0],x3.shape[1]+1))
x4[:,:-1]=x3
x4[:,-1]=cloud

x5=np.hstack((x4,m))

x6=np.zeros((x5.shape[0],x5.shape[1]+1))
x6[:,:-1]=x5
x6[:,-1]=year

x7=np.zeros((x6.shape[0],x6.shape[1]+1))
x7[:,:-1]=x6
x7[:,-1]=night


#x8=np.zeros((x7.shape[0],x7.shape[1]+1))
#x8[:,:-1]=x7
#x8[:,-1]=closed

x9=np.zeros((x7.shape[0],x7.shape[1]+1))
x9[:,:-1]=x7
x9[:,-1]= h

x_test=x9
#x=np.hstack((x9,date))

#x_test = [[0,	0,	0,	0,	1,	0,	0,	-0.879784,	-0.245839,	0,	1.15995,	0,	0,	0,	0,	0,	0,	0,	0,	0,	0,	0,	1,	0.7,	0	,-0.230377]]

y_predict = model.predict(x_test)
print(y_predict)